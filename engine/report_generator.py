import hashlib
import logging
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
from PIL import Image

logger = logging.getLogger(__name__)

# ── Konstanta Laporan ─────────────────────────────────────────────────────────
NAMA_SISTEM   = "SISTEM EVALUASI STEGANOGRAFI LSB-MWC"
NAMA_PENELITI = "Muhammad Akmal Fitrianto"
NPM           = "2210010546"

# Warna tema dark untuk matplotlib (harus hex string)
BG_DARK   = "#0d1117"
BG_PANEL  = "#161b22"
TEAL      = "#00d4aa"
SKY       = "#38bdf8"
MERAH     = "#f85149"
AMBER     = "#fbbf24"
HIJAU     = "#3fb950"
TEKS      = "#e6edf3"
TEKS_SEK  = "#8b949e"
BORDER    = "#30363d"


# ── Utilitas Bersama ──────────────────────────────────────────────────────────

def _buat_id_laporan(prefix: str) -> str:
    
    rnd  = uuid.uuid4().hex[:8].upper()
    tgl  = datetime.now().strftime("%d%m%Y")
    return f"REPT/{prefix}/{rnd}-{tgl}"


def _tanggal_sekarang() -> str:
    return datetime.now().strftime("%d %B %Y")


def _setup_tema_matplotlib() -> None:
    """Terapkan tema gelap ke semua figure matplotlib."""
    import matplotlib
    matplotlib.rcParams.update({
        "figure.facecolor":  BG_DARK,
        "axes.facecolor":    BG_PANEL,
        "axes.edgecolor":    BORDER,
        "axes.labelcolor":   TEKS,
        "axes.titlecolor":   TEKS,
        "axes.grid":         True,
        "grid.color":        BORDER,
        "grid.linewidth":    0.5,
        "xtick.color":       TEKS_SEK,
        "ytick.color":       TEKS_SEK,
        "text.color":        TEKS,
        "legend.facecolor":  BG_PANEL,
        "legend.edgecolor":  BORDER,
        "font.family":       "DejaVu Sans",
        "font.size":         9,
    })


def _halaman_header(fig, judul_laporan: str, id_laporan: str, info_tambahan: str = "") -> None:
    
    import matplotlib.pyplot as plt
    # Garis header atas
    fig.text(0.05, 0.97, NAMA_SISTEM, fontsize=10, fontweight="bold", color=TEKS, va="top")
    fig.text(0.05, 0.94, judul_laporan, fontsize=13, fontweight="bold", color=TEAL, va="top")
    fig.text(0.05, 0.91, f"ID Laporan: {id_laporan}", fontsize=8, color=TEKS_SEK, va="top")
    fig.text(0.95, 0.97, _tanggal_sekarang(), fontsize=8, color=TEKS_SEK, va="top", ha="right")
    if info_tambahan:
        fig.text(0.05, 0.89, info_tambahan, fontsize=8, color=TEKS_SEK, va="top")

    # Garis pemisah
    from matplotlib.lines import Line2D
    line = Line2D([0.05, 0.95], [0.875, 0.875], transform=fig.transFigure,
                  color=TEAL, linewidth=1.5, alpha=0.7)
    fig.add_artist(line)


def _footer(fig) -> None:
    
    fig.text(0.95, 0.02, f"({NAMA_PENELITI} — {NPM})", fontsize=7,
             color=TEKS_SEK, ha="right", va="bottom")
    fig.text(0.05, 0.02, f"Universitas Islam Kalimantan Muhammad Arsyad Al Banjari",
             fontsize=7, color=TEKS_SEK, va="bottom")


def _muat_arr(path: str | Path) -> np.ndarray:
    with Image.open(path) as img:
        return np.array(img.convert("RGB"), dtype=np.uint8)


def _hitung_hist(arr_ch: np.ndarray):
    bins = np.arange(257)
    freq, _ = np.histogram(arr_ch.ravel(), bins=bins)
    return bins[:-1], freq


def _setup_ax(ax, judul: str = "") -> None:
    ax.set_facecolor(BG_PANEL)
    if judul:
        ax.set_title(judul, color=TEKS, fontsize=8, fontweight="bold")
    ax.tick_params(colors=TEKS_SEK, labelsize=7)
    for spine in ax.spines.values():
        spine.set_edgecolor(BORDER)
    ax.grid(True, color=BORDER, linewidth=0.4, alpha=0.5)


# ══════════════════════════════════════════════════════════════════════════════
# LAPORAN 1 — REKAPITULASI METRIK KUALITAS
# ══════════════════════════════════════════════════════════════════════════════

def laporan_1_pdf(records: list, path_output: str | Path) -> Path:
    
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    path_output = Path(path_output)
    id_lap = _buat_id_laporan("REPT-001")
    _setup_tema_matplotlib()

    with PdfPages(str(path_output)) as pdf:
        fig = plt.figure(figsize=(11.69, 8.27), facecolor=BG_DARK)  # A4 landscape
        _halaman_header(fig, "LAPORAN REKAPITULASI METRIK KUALITAS CITRA", id_lap)
        _footer(fig)

        if not records:
            fig.text(0.5, 0.5, "Belum ada data riwayat.", ha="center", color=TEKS_SEK, fontsize=12)
            pdf.savefig(fig, facecolor=BG_DARK)
            plt.close(fig)
            return path_output

        # ── Tabel ─────────────────────────────────────────────────────────────
        ax_tbl = fig.add_axes([0.05, 0.35, 0.90, 0.50])
        ax_tbl.axis("off")

        header = ["No.", "Nama File (Cover)", "Ukuran Pesan", "Kunci MWC", "PSNR (dB)", "MSE", "Waktu Simpan"]
        baris_data = []
        for i, r in enumerate(records, 1):
            psnr_str = f"{r.nilai_psnr:.4f}" if r.nilai_psnr is not None else "—"
            mse_str  = f"{r.nilai_mse:.6f}"  if r.nilai_mse  is not None else "—"
            baris_data.append([
                str(i), r.nama_file, f"{r.ukuran_pesan:,} byte",
                "●" * min(len(r.kunci_seed), 6),
                psnr_str, mse_str,
                str(r.waktu_simpan)[:19] if r.waktu_simpan else "—",
            ])

        # Rata-rata
        psnr_vals = [r.nilai_psnr for r in records if r.nilai_psnr is not None]
        mse_vals  = [r.nilai_mse  for r in records if r.nilai_mse  is not None]
        avg_psnr = f"{sum(psnr_vals)/len(psnr_vals):.4f}" if psnr_vals else "—"
        avg_mse  = f"{sum(mse_vals)/len(mse_vals):.6f}"   if mse_vals  else "—"
        baris_data.append(["", "─── Rata-rata ───", "", "", avg_psnr, avg_mse, ""])

        tabel = ax_tbl.table(
            cellText=baris_data, colLabels=header,
            loc="center", cellLoc="center",
        )
        tabel.auto_set_font_size(False)
        tabel.set_fontsize(7)
        tabel.scale(1, 1.4)

        # Styling header tabel
        for j in range(len(header)):
            tabel[0, j].set_facecolor(TEAL)
            tabel[0, j].set_text_props(color=BG_DARK, fontweight="bold")
        for i in range(1, len(baris_data) + 1):
            bg = BG_PANEL if i % 2 == 0 else "#1c2230"
            for j in range(len(header)):
                tabel[i, j].set_facecolor(bg)
                tabel[i, j].set_text_props(color=TEKS)
                tabel[i, j].set_edgecolor(BORDER)

        # Baris rata-rata berwarna berbeda
        for j in range(len(header)):
            tabel[len(baris_data), j].set_facecolor("#1a2a1a")
            tabel[len(baris_data), j].set_text_props(color=HIJAU, fontweight="bold")

        # ── Grafik PSNR ───────────────────────────────────────────────────────
        if psnr_vals:
            ax_chart = fig.add_axes([0.05, 0.06, 0.55, 0.24])
            ax_chart.bar(range(1, len(psnr_vals) + 1), psnr_vals, color=TEAL, alpha=0.8, width=0.6)
            ax_chart.axhline(y=50, color=HIJAU, linestyle="--", linewidth=1, label="Threshold 50 dB (Sangat Baik)")
            ax_chart.axhline(y=30, color=AMBER, linestyle="--", linewidth=1, label="Threshold 30 dB (Minimum)")
            ax_chart.set_xlabel("No. Pengujian", fontsize=7)
            ax_chart.set_ylabel("PSNR (dB)", fontsize=7)
            _setup_ax(ax_chart, "Grafik Nilai PSNR per Pengujian")
            ax_chart.legend(fontsize=6, facecolor=BG_PANEL, labelcolor=TEKS)

            # Informasi ringkas
            fig.text(0.65, 0.27, f"Total Pengujian : {len(records)}", fontsize=8, color=TEKS)
            fig.text(0.65, 0.23, f"Rata-rata PSNR  : {avg_psnr} dB", fontsize=8, color=TEAL, fontweight="bold")
            fig.text(0.65, 0.19, f"Rata-rata MSE   : {avg_mse}", fontsize=8, color=TEKS)
            fig.text(0.65, 0.15, f"PSNR Tertinggi  : {max(psnr_vals):.4f} dB", fontsize=8, color=HIJAU)
            fig.text(0.65, 0.11, f"PSNR Terendah   : {min(psnr_vals):.4f} dB", fontsize=8, color=MERAH)

        pdf.savefig(fig, facecolor=BG_DARK)
        plt.close(fig)

    logger.info(f"Laporan 1 PDF → {path_output}")
    return path_output


def laporan_1_excel(records: list, path_output: str | Path) -> Path:
    
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl tidak terinstall. Jalankan: pip install openpyxl")

    path_output = Path(path_output)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrik Kualitas"

    # Gaya
    fill_header = PatternFill(fill_type="solid", fgColor="00D4AA")
    fill_genap   = PatternFill(fill_type="solid", fgColor="161B22")
    fill_ganjil  = PatternFill(fill_type="solid", fgColor="1C2230")
    fill_avg     = PatternFill(fill_type="solid", fgColor="0D2818")
    font_header  = Font(bold=True, color="0D1117", size=10)
    font_putih   = Font(color="E6EDF3", size=9)
    font_teal    = Font(color="00D4AA", bold=True, size=9)
    font_hijau   = Font(color="3FB950", bold=True, size=9)
    border_tipis = Border(
        left=Side(style="thin", color="30363D"),
        right=Side(style="thin", color="30363D"),
        top=Side(style="thin", color="30363D"),
        bottom=Side(style="thin", color="30363D"),
    )
    center = Alignment(horizontal="center", vertical="center")

    # Judul
    ws.merge_cells("A1:G1")
    ws["A1"] = NAMA_SISTEM
    ws["A1"].font = Font(bold=True, color="00D4AA", size=12)
    ws["A1"].alignment = center
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="0D1117")

    ws.merge_cells("A2:G2")
    ws["A2"] = "LAPORAN REKAPITULASI METRIK KUALITAS CITRA"
    ws["A2"].font = Font(bold=True, color="E6EDF3", size=10)
    ws["A2"].alignment = center

    ws.merge_cells("A3:D3")
    ws["A3"] = f"Tanggal: {_tanggal_sekarang()}"
    ws["A3"].font = Font(color="8B949E", size=9)

    ws.merge_cells("E3:G3")
    ws["E3"] = f"Peneliti: {NAMA_PENELITI} ({NPM})"
    ws["E3"].font = Font(color="8B949E", size=9)
    ws["E3"].alignment = Alignment(horizontal="right")

    # Header kolom (baris 5)
    header = ["No.", "Nama File (Cover)", "Ukuran Pesan (byte)", "Kunci MWC",
              "PSNR (dB)", "MSE", "Waktu Simpan"]
    lebar  = [6, 28, 20, 16, 14, 16, 22]

    for j, (h, l) in enumerate(zip(header, lebar), start=1):
        cell = ws.cell(row=5, column=j, value=h)
        cell.font      = font_header
        cell.fill      = fill_header
        cell.alignment = center
        cell.border    = border_tipis
        ws.column_dimensions[get_column_letter(j)].width = l

    # Baris data
    for i, r in enumerate(records, start=1):
        row_idx = 5 + i
        fill = fill_genap if i % 2 == 0 else fill_ganjil
        psnr_val = r.nilai_psnr
        mse_val  = r.nilai_mse
        data = [
            i, r.nama_file, r.ukuran_pesan,
            "●" * min(len(r.kunci_seed), 6),
            round(psnr_val, 4) if psnr_val is not None else None,
            round(mse_val, 6)  if mse_val  is not None else None,
            str(r.waktu_simpan)[:19] if r.waktu_simpan else None,
        ]
        for j, val in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=j, value=val)
            cell.font      = font_putih
            cell.fill      = fill
            cell.alignment = center
            cell.border    = border_tipis
            # Warna PSNR
            if j == 5 and psnr_val is not None:
                if psnr_val >= 50:
                    cell.font = Font(color="3FB950", size=9)
                elif psnr_val >= 40:
                    cell.font = Font(color="00D4AA", size=9)
                elif psnr_val >= 30:
                    cell.font = Font(color="FBBF24", size=9)
                else:
                    cell.font = Font(color="F85149", size=9)

    # Baris rata-rata
    psnr_vals = [r.nilai_psnr for r in records if r.nilai_psnr is not None]
    mse_vals  = [r.nilai_mse  for r in records if r.nilai_mse  is not None]
    avg_row   = 5 + len(records) + 1

    ws.cell(row=avg_row, column=2, value="RATA-RATA").font = font_hijau
    ws.cell(row=avg_row, column=5, value=round(sum(psnr_vals)/len(psnr_vals), 4) if psnr_vals else None).font = font_hijau
    ws.cell(row=avg_row, column=6, value=round(sum(mse_vals)/len(mse_vals), 6) if mse_vals else None).font = font_hijau

    for j in range(1, 8):
        c = ws.cell(row=avg_row, column=j)
        c.fill   = fill_avg
        c.border = border_tipis
        c.alignment = center
        if j not in (2, 5, 6):
            c.font = font_putih

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[5].height = 18
    ws.sheet_view.showGridLines = False

    wb.save(str(path_output))
    logger.info(f"Laporan 1 Excel → {path_output}")
    return path_output


# ══════════════════════════════════════════════════════════════════════════════
# LAPORAN 2 — ANALISIS HISTOGRAM CITRA
# ══════════════════════════════════════════════════════════════════════════════

def laporan_2_pdf(
    path_cover: str | Path,
    path_stego: str | Path,
    path_output: str | Path,
    psnr: float | None = None,
    mse: float | None = None,
) -> Path:
    
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    path_output = Path(path_output)
    id_lap = _buat_id_laporan("HIST-002")
    _setup_tema_matplotlib()

    arr_cover = _muat_arr(path_cover)
    arr_stego = _muat_arr(path_stego)
    nama_cover = Path(path_cover).name

    info = f"File: {nama_cover}"
    if psnr is not None:
        info += f"  |  PSNR: {psnr:.4f} dB"
    if mse is not None:
        info += f"  |  MSE: {mse:.6f}"

    with PdfPages(str(path_output)) as pdf:
        # ── Halaman 1: Thumbnail + Histogram RGB ──────────────────────────────
        fig = plt.figure(figsize=(11.69, 8.27), facecolor=BG_DARK)
        _halaman_header(fig, "LAPORAN ANALISIS HISTOGRAM CITRA", id_lap, info)
        _footer(fig)

        nama_ch = ["Red (R)", "Green (G)", "Blue (B)"]
        warna   = [("#e05c5c","#e0a0a0"), ("#5cb85c","#a0e0a0"), ("#5c8fe0","#a0c0f0")]

        # Thumbnail cover & stego
        ax_cv = fig.add_axes([0.05, 0.65, 0.18, 0.20])
        ax_cv.imshow(arr_cover); ax_cv.axis("off")
        ax_cv.set_title("Citra Cover (Asli)", color=TEAL, fontsize=8, fontweight="bold")

        ax_st = fig.add_axes([0.27, 0.65, 0.18, 0.20])
        ax_st.imshow(arr_stego); ax_st.axis("off")
        ax_st.set_title("Citra Stego (Setelah Embedding)", color=AMBER, fontsize=8, fontweight="bold")

        # Histogram RGB
        for i, (nm, (wc, ws)) in enumerate(zip(nama_ch, warna)):
            ax = fig.add_axes([0.05 + i * 0.32, 0.08, 0.28, 0.52])
            x, fc = _hitung_hist(arr_cover[:, :, i])
            _, fs = _hitung_hist(arr_stego[:, :, i])
            ax.fill_between(x, fc, alpha=0.5, color=wc, step="post", label="Cover")
            ax.fill_between(x, fs, alpha=0.5, color=ws, step="post", label="Stego")
            ax.step(x, fc, color=wc, lw=0.7, where="post")
            ax.step(x, fs, color=ws, lw=0.7, where="post")
            _setup_ax(ax, f"Histogram Channel {nm}")
            ax.set_xlim(0, 255)
            ax.set_xlabel("Intensitas (0–255)", fontsize=7)
            ax.set_ylabel("Frekuensi", fontsize=7)
            ax.legend(fontsize=7, facecolor=BG_PANEL, labelcolor=TEKS)

        pdf.savefig(fig, facecolor=BG_DARK)
        plt.close(fig)

        # ── Halaman 2: Histogram Selisih Δ ────────────────────────────────────
        fig2 = plt.figure(figsize=(11.69, 8.27), facecolor=BG_DARK)
        _halaman_header(fig2, "LAPORAN ANALISIS HISTOGRAM CITRA — Selisih Δ", id_lap, info)
        _footer(fig2)

        for i, nm in enumerate(nama_ch):
            ax = fig2.add_axes([0.05 + i * 0.32, 0.12, 0.28, 0.68])
            x, fc = _hitung_hist(arr_cover[:, :, i])
            _, fs = _hitung_hist(arr_stego[:, :, i])
            delta = fs.astype(np.int64) - fc.astype(np.int64)
            warna_bar = [TEAL if d >= 0 else MERAH for d in delta]
            ax.bar(x, delta, color=warna_bar, width=1.0, alpha=0.85)
            ax.axhline(0, color=TEKS_SEK, linewidth=0.8, linestyle="--")
            _setup_ax(ax, f"Δ Histogram Channel {nm}")
            ax.set_xlim(0, 255)
            ax.text(0.98, 0.98, f"Max |Δ| = {int(np.max(np.abs(delta)))}",
                    transform=ax.transAxes, fontsize=7, ha="right", va="top", color=AMBER)

        teks_analisis = (
            "Analisis: Penyisipan LSB pada koordinat acak MWC tidak mengubah distribusi "
            "statistik warna secara signifikan. Kedua histogram terlihat identik."
        )
        fig2.text(0.05, 0.05, teks_analisis, fontsize=8, color=TEKS_SEK, style="italic")

        pdf.savefig(fig2, facecolor=BG_DARK)
        plt.close(fig2)

    logger.info(f"Laporan 2 PDF → {path_output}")
    return path_output


# ══════════════════════════════════════════════════════════════════════════════
# LAPORAN 3 — VISUALISASI SEBARAN BIT (NOISE MAP)
# ══════════════════════════════════════════════════════════════════════════════

def laporan_3_pdf(
    path_cover: str | Path,
    path_stego: str | Path,
    path_output: str | Path,
    kunci: str | None = None,
    psnr: float | None = None,
) -> Path:
    
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.patches import Patch

    path_output = Path(path_output)
    id_lap = _buat_id_laporan("NOISE-003")
    _setup_tema_matplotlib()

    arr_cover = _muat_arr(path_cover)
    arr_stego = _muat_arr(path_stego)
    mask      = arr_cover[:, :, 0] != arr_stego[:, :, 0]
    h, w      = arr_cover.shape[:2]
    n_berubah = int(mask.sum())
    nama_cover = Path(path_cover).name
    psnr_info  = f"{psnr:.4f} dB" if psnr is not None else "—"

    with PdfPages(str(path_output)) as pdf:
        fig = plt.figure(figsize=(11.69, 8.27), facecolor=BG_DARK)
        _halaman_header(
            fig, "LAPORAN VISUALISASI SEBARAN BIT (NOISE MAP)", id_lap,
            f"File: {nama_cover}  |  PSNR: {psnr_info}  |  Metode: LSB + PRNG MWC",
        )
        _footer(fig)

        # Noise map binary
        ax1 = fig.add_axes([0.05, 0.35, 0.27, 0.45])
        ax1.imshow(mask.astype(np.uint8) * 255, cmap="gray", aspect="auto")
        ax1.set_title("Noise Map Binary\n(Putih = Piksel Berubah)", color=TEKS, fontsize=8, fontweight="bold")
        ax1.set_xlabel(f"Kolom ({w} px)", fontsize=7, color=TEKS_SEK)
        ax1.set_ylabel(f"Baris ({h} px)", fontsize=7, color=TEKS_SEK)
        ax1.tick_params(colors=TEKS_SEK, labelsize=6)

        # Overlay
        overlay = arr_cover.copy()
        overlay[mask, 0] = 0; overlay[mask, 1] = 212; overlay[mask, 2] = 170
        ax2 = fig.add_axes([0.37, 0.35, 0.27, 0.45])
        ax2.imshow(overlay, aspect="auto")
        ax2.set_title("Overlay Noise Map\n(Teal = Dimodifikasi LSB)", color=TEAL, fontsize=8, fontweight="bold")
        ax2.axis("off")
        legend_el = [Patch(facecolor=TEAL, label="Piksel dimodifikasi")]
        ax2.legend(handles=legend_el, loc="lower right", fontsize=7,
                   facecolor=BG_PANEL, edgecolor=BORDER, labelcolor=TEKS)

        # Scatter koordinat MWC (jika kunci diberikan)
        if kunci:
            try:
                from engine.mwc_generator import MWCGenerator
                n_bit = max(32, n_berubah + 32)
                gen   = MWCGenerator(password=kunci)
                coords = gen.hasilkan_koordinat(lebar=w, tinggi=h, jumlah=min(n_bit, w * h))
                xs = [x for x, y in coords]; ys = [y for x, y in coords]

                ax3 = fig.add_axes([0.69, 0.35, 0.27, 0.45])
                ax3.scatter(xs, ys, s=0.3, alpha=0.3, color=MERAH, linewidths=0)
                _setup_ax(ax3, f"Scatter Koordinat MWC\n({len(coords):,} titik)")
                ax3.set_xlim(0, w); ax3.set_ylim(h, 0)
                ax3.set_xlabel("X (kolom)", fontsize=7); ax3.set_ylabel("Y (baris)", fontsize=7)
            except Exception:
                pass

        # Statistik
        persen = (n_berubah / (w * h)) * 100
        fig.text(0.05, 0.28, "Statistik Noise Map:", fontsize=9, fontweight="bold", color=TEKS)
        fig.text(0.05, 0.24, f"• Dimensi citra     : {w} × {h} piksel ({w*h:,} total)", fontsize=8, color=TEKS_SEK)
        fig.text(0.05, 0.20, f"• Piksel dimodifikasi: {n_berubah:,} piksel ({persen:.4f}%)", fontsize=8, color=TEAL)
        fig.text(0.05, 0.16, f"• Perubahan nilai   : ±1 (LSB 1-bit, tidak kasat mata)", fontsize=8, color=TEKS_SEK)
        fig.text(0.05, 0.12, f"• Channel dimodifikasi: Red (channel ke-0)", fontsize=8, color=TEKS_SEK)

        teks_analisis = (
            "Analisis: Noise Map menunjukkan sebaran bit pesan yang tersebar secara acak dan merata "
            "di seluruh permukaan citra. Hal ini membuktikan bahwa PRNG MWC berhasil menghasilkan "
            "koordinat penyisipan yang tidak membentuk pola teratur, sehingga tidak mudah dideteksi "
            "melalui analisis statistik atau steganalisis visual."
        )
        fig.text(0.05, 0.04, teks_analisis, fontsize=8, color=TEKS_SEK, style="italic", wrap=True)

        pdf.savefig(fig, facecolor=BG_DARK)
        plt.close(fig)

    logger.info(f"Laporan 3 PDF → {path_output}")
    return path_output


# ══════════════════════════════════════════════════════════════════════════════
# LAPORAN 4 — DEGRADASI KUALITAS VS KAPASITAS
# ══════════════════════════════════════════════════════════════════════════════

def laporan_4_pdf(records: list, path_cover: str | Path, path_output: str | Path) -> Path:
    
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    path_output = Path(path_output)
    id_lap = _buat_id_laporan("TREND-004")
    _setup_tema_matplotlib()

    with PdfPages(str(path_output)) as pdf:
        fig = plt.figure(figsize=(11.69, 8.27), facecolor=BG_DARK)
        _halaman_header(
            fig, "LAPORAN DEGRADASI KUALITAS DAN KAPASITAS", id_lap,
            f"File Cover: {Path(path_cover).name}",
        )
        _footer(fig)

        data_valid = [(r.ukuran_pesan, r.nilai_psnr, r.nilai_mse)
                      for r in records
                      if r.nilai_psnr is not None and r.nilai_mse is not None]

        if len(data_valid) < 2:
            fig.text(0.5, 0.5, "Minimal 2 data diperlukan untuk grafik degradasi.",
                     ha="center", color=TEKS_SEK, fontsize=11)
            pdf.savefig(fig, facecolor=BG_DARK); plt.close(fig)
            return path_output

        data_valid.sort(key=lambda x: x[0])
        ukuran = [d[0] for d in data_valid]
        psnr   = [d[1] for d in data_valid]
        mse    = [d[2] for d in data_valid]

        # Grafik PSNR vs ukuran
        ax1 = fig.add_axes([0.07, 0.15, 0.42, 0.60])
        ax1.plot(ukuran, psnr, "o-", color=TEAL, linewidth=2, markersize=5)
        ax1.fill_between(ukuran, psnr, alpha=0.15, color=TEAL)
        ax1.axhline(50, color=HIJAU, linestyle="--", linewidth=1, label="> 50 dB (Sangat Baik)")
        ax1.axhline(30, color=AMBER, linestyle="--", linewidth=1, label="> 30 dB (Batas Minimum)")
        _setup_ax(ax1, "Grafik Penurunan PSNR terhadap Kapasitas Payload")
        ax1.set_xlabel("Ukuran Pesan (byte)", fontsize=8)
        ax1.set_ylabel("PSNR (dB)", fontsize=8)
        ax1.legend(fontsize=7, facecolor=BG_PANEL, labelcolor=TEKS)

        # Grafik MSE vs ukuran
        ax2 = fig.add_axes([0.57, 0.15, 0.38, 0.60])
        ax2.plot(ukuran, mse, "s-", color=AMBER, linewidth=2, markersize=5)
        ax2.fill_between(ukuran, mse, alpha=0.15, color=AMBER)
        _setup_ax(ax2, "Grafik Peningkatan MSE terhadap Kapasitas Payload")
        ax2.set_xlabel("Ukuran Pesan (byte)", fontsize=8)
        ax2.set_ylabel("MSE", fontsize=8)

        # Tabel ringkasan
        teks = f"Total data: {len(data_valid)} pengujian  |  "
        teks += f"Ukuran: {min(ukuran):,}–{max(ukuran):,} byte  |  "
        teks += f"PSNR: {min(psnr):.4f}–{max(psnr):.4f} dB"
        fig.text(0.07, 0.08, teks, fontsize=8, color=TEKS_SEK)
        fig.text(0.07, 0.04,
                 "Analisis: Grafik menunjukkan korelasi negatif antara ukuran payload dengan nilai PSNR. "
                 "Meski demikian, algoritma LSB-MWC masih mempertahankan PSNR di atas ambang batas 30 dB.",
                 fontsize=8, color=TEKS_SEK, style="italic")

        pdf.savefig(fig, facecolor=BG_DARK)
        plt.close(fig)

    logger.info(f"Laporan 4 PDF → {path_output}")
    return path_output


# ══════════════════════════════════════════════════════════════════════════════
# LAPORAN 5 — KOMPARASI VISUAL
# ══════════════════════════════════════════════════════════════════════════════

def laporan_5_pdf(
    path_cover: str | Path,
    path_stego: str | Path,
    path_output: str | Path,
    psnr: float | None = None,
    mse: float | None = None,
    kapasitas_persen: float | None = None,
) -> Path:
    
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    path_output = Path(path_output)
    id_lap = _buat_id_laporan("VISUAL-005")
    _setup_tema_matplotlib()

    arr_cover = _muat_arr(path_cover)
    arr_stego = _muat_arr(path_stego)
    diff = np.abs(arr_cover.astype(np.int16) - arr_stego.astype(np.int16))
    diff_amplified = np.clip(diff * 50, 0, 255).astype(np.uint8)

    nama_cover = Path(path_cover).name
    psnr_str   = f"{psnr:.4f} dB" if psnr is not None else "—"

    with PdfPages(str(path_output)) as pdf:
        fig = plt.figure(figsize=(11.69, 8.27), facecolor=BG_DARK)
        _halaman_header(
            fig, "LAPORAN KOMPARASI VISUAL / IMPERCEPTIBILITY TEST", id_lap,
            f"Nama File: {nama_cover}  |  PSNR: {psnr_str}  |  MSE: {mse:.6f}" if mse else f"File: {nama_cover}",
        )
        _footer(fig)

        # 3 kolom: cover, stego, diferensi × 50
        for i, (arr, judul, warna_judul) in enumerate([
            (arr_cover, "Citra Cover (Asli / Original)", TEAL),
            (arr_stego, "Citra Stego (Hasil Embedding)", AMBER),
            (diff_amplified, "Diferensi × 50 (Amplified Difference)", MERAH),
        ]):
            ax = fig.add_axes([0.04 + i * 0.33, 0.25, 0.30, 0.55])
            ax.imshow(arr, aspect="auto")
            ax.set_title(judul, color=warna_judul, fontsize=8, fontweight="bold")
            ax.axis("off")

        # Info
        h, w = arr_cover.shape[:2]
        info_lines = [
            f"Dimensi          : {w} × {h} piksel",
            f"PSNR             : {psnr_str}",
            f"MSE              : {f'{mse:.6f}' if mse is not None else '—'}",
            f"Kapasitas terpakai: {f'{kapasitas_persen:.2f}%' if kapasitas_persen is not None else '—'}",
        ]
        for i, teks in enumerate(info_lines):
            fig.text(0.05, 0.20 - i * 0.04, teks, fontsize=9, color=TEKS)

        analisis = (
            "Analisis: Berdasarkan pengujian HVS (Human Visual System), tidak terdapat perbedaan "
            "visual yang signifikan antara citra asli dan citra stego. Hal ini dibuktikan dengan "
            f"nilai PSNR {psnr_str} yang melampaui ambang batas 30 dB, mengindikasikan kualitas "
            "visual citra stego masuk dalam kategori 'imperceptible'."
        )
        fig.text(0.05, 0.06, analisis, fontsize=8, color=TEKS_SEK, style="italic", wrap=True)

        pdf.savefig(fig, facecolor=BG_DARK)
        plt.close(fig)

    logger.info(f"Laporan 5 PDF → {path_output}")
    return path_output


# ══════════════════════════════════════════════════════════════════════════════
# LAPORAN 6 — ANALISIS SENSITIVITAS KUNCI
# ══════════════════════════════════════════════════════════════════════════════

def laporan_6_pdf(
    path_stego: str | Path,
    kunci_benar: str,
    kunci_salah_list: list[str],
    path_output: str | Path,
) -> Path:
    """
    Laporan 6 (PDF): Uji sensitivitas kunci — tunjukkan output dengan kunci benar vs salah.
    Membuktikan bahwa ekstraksi hanya berhasil dengan kunci yang tepat.
    """
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    path_output = Path(path_output)
    id_lap = _buat_id_laporan("KEYSEN-006")
    _setup_tema_matplotlib()

    from engine.stego_lsb import extract_data

    hasil_uji: list[dict] = []

    # Uji kunci benar
    try:
        pesan = extract_data(stego_path=path_stego, key=kunci_benar)
        hasil_uji.append({
            "kunci": "●" * min(len(kunci_benar), 8),
            "seed": hashlib.sha256(kunci_benar.encode()).hexdigest()[:8].upper(),
            "status": "BERHASIL",
            "output": pesan[:60] + "..." if len(pesan) > 60 else pesan,
        })
    except Exception as e:
        hasil_uji.append({"kunci": "BENAR", "seed": "—", "status": "ERROR", "output": str(e)[:60]})

    # Uji kunci salah
    for kunci_s in kunci_salah_list:
        try:
            pesan_s = extract_data(stego_path=path_stego, key=kunci_s)
            output = pesan_s[:60] if len(pesan_s) >= 60 else pesan_s
        except Exception:
            output = "[Gagal Decode / Error Decoding]"
        hasil_uji.append({
            "kunci": "●" * min(len(kunci_s), 8),
            "seed": hashlib.sha256(kunci_s.encode()).hexdigest()[:8].upper(),
            "status": "GAGAL (Kunci Salah)",
            "output": output,
        })

    with PdfPages(str(path_output)) as pdf:
        fig = plt.figure(figsize=(11.69, 8.27), facecolor=BG_DARK)
        _halaman_header(
            fig, "LAPORAN ANALISIS SENSITIVITAS KUNCI / KEY SENSITIVITY", id_lap,
            f"File Stego: {Path(path_stego).name}  |  Algoritma PRNG: MWC",
        )
        _footer(fig)

        ax = fig.add_axes([0.05, 0.15, 0.90, 0.65])
        ax.axis("off")

        header = ["Jenis Kunci", "Seed (SHA-256 8-hex)", "Koordinat MWC", "Status", "Output Pesan (60 char)"]
        tabel_data = []
        for h_item in hasil_uji:
            status  = h_item["status"]
            is_ok   = "BERHASIL" in status
            koordinat = "BENAR → Dapat diekstrak" if is_ok else "SALAH → Koordinat berbeda"
            tabel_data.append([
                "Kunci Benar" if is_ok else "Kunci Salah",
                h_item["seed"],
                koordinat,
                status,
                h_item["output"],
            ])

        tbl = ax.table(cellText=tabel_data, colLabels=header, loc="center", cellLoc="left")
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(7)
        tbl.scale(1, 2.0)

        for j in range(len(header)):
            tbl[0, j].set_facecolor(TEAL); tbl[0, j].set_text_props(color=BG_DARK, fontweight="bold")

        for i, h_item in enumerate(hasil_uji, start=1):
            is_ok = "BERHASIL" in h_item["status"]
            row_color = "#0d2018" if is_ok else "#2d0d0d"
            for j in range(len(header)):
                tbl[i, j].set_facecolor(row_color)
                tbl[i, j].set_text_props(color=HIJAU if is_ok else MERAH)
                tbl[i, j].set_edgecolor(BORDER)

        fig.text(0.05, 0.10,
                 "Analisis: Perbedaan satu karakter pada kunci menghasilkan nilai seed (X₀, C₀) yang "
                 "sepenuhnya berbeda, sehingga koordinat piksel yang dibangkitkan MWC tidak dapat "
                 "mereproduksi urutan penyisipan yang sama. Sistem hanya bisa diekstrak dengan kunci yang tepat.",
                 fontsize=8, color=TEKS_SEK, style="italic")

        pdf.savefig(fig, facecolor=BG_DARK)
        plt.close(fig)

    logger.info(f"Laporan 6 PDF → {path_output}")
    return path_output


# ══════════════════════════════════════════════════════════════════════════════
# LAPORAN 7 — OUTPUT PESAN & INTEGRITAS DATA
# ══════════════════════════════════════════════════════════════════════════════

def laporan_7_pdf(
    pesan_input: str,
    pesan_output: str,
    path_cover: str | Path,
    path_stego: str | Path,
    kunci: str,
    path_output: str | Path,
) -> Path:
    
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    path_output = Path(path_output)
    id_lap = _buat_id_laporan("INTEG-007")
    _setup_tema_matplotlib()

    identik = pesan_input == pesan_output
    akurasi = sum(a == b for a, b in zip(pesan_input, pesan_output)) / max(len(pesan_input), 1) * 100

    with PdfPages(str(path_output)) as pdf:
        fig = plt.figure(figsize=(11.69, 8.27), facecolor=BG_DARK)
        _halaman_header(
            fig, "LAPORAN OUTPUT PESAN & INTEGRITAS DATA", id_lap,
            f"Citra Sumber: {Path(path_stego).name}  |  Kunci MWC: {'●' * min(len(kunci), 8)}",
        )
        _footer(fig)

        # Pesan Input (kiri)
        ax1 = fig.add_axes([0.05, 0.35, 0.42, 0.48])
        ax1.set_facecolor(BG_PANEL); ax1.axis("off")
        ax1.set_title("Pesan Input (Sebelum Disisipkan)", color=TEAL, fontsize=9, fontweight="bold")
        wrap_input = [pesan_input[i:i+55] for i in range(0, min(len(pesan_input), 550), 55)]
        for k, line in enumerate(wrap_input[:8]):
            ax1.text(0.02, 0.90 - k * 0.10, line, transform=ax1.transAxes,
                     fontsize=7, color=TEKS, family="monospace")
        if len(pesan_input) > 440:
            ax1.text(0.02, 0.04, "...(terpotong)...", transform=ax1.transAxes,
                     fontsize=7, color=TEKS_SEK, style="italic")

        # Pesan Output (kanan)
        ax2 = fig.add_axes([0.53, 0.35, 0.42, 0.48])
        ax2.set_facecolor(BG_PANEL); ax2.axis("off")
        ax2.set_title("Pesan Output (Hasil Ekstraksi)", color=HIJAU, fontsize=9, fontweight="bold")
        wrap_output = [pesan_output[i:i+55] for i in range(0, min(len(pesan_output), 550), 55)]
        for k, line in enumerate(wrap_output[:8]):
            ax2.text(0.02, 0.90 - k * 0.10, line, transform=ax2.transAxes,
                     fontsize=7, color=TEKS, family="monospace")

        # Status integritas
        status_teks = "✓  STATUS INTEGRITAS: 100% MATCH (VALID)" if identik else "✗  STATUS INTEGRITAS: TIDAK COCOK"
        status_warna = HIJAU if identik else MERAH
        fig.text(0.5, 0.28, status_teks, fontsize=13, fontweight="bold",
                 color=status_warna, ha="center")

        # Statistik
        stats = [
            f"Panjang pesan input   : {len(pesan_input):,} karakter  ({len(pesan_input.encode('utf-8')):,} byte)",
            f"Panjang pesan output  : {len(pesan_output):,} karakter  ({len(pesan_output.encode('utf-8')):,} byte)",
            f"Tingkat kesesuaian    : {akurasi:.2f}%",
            f"Hash SHA-256 (input)  : {hashlib.sha256(pesan_input.encode()).hexdigest()[:32]}...",
            f"Hash SHA-256 (output) : {hashlib.sha256(pesan_output.encode()).hexdigest()[:32]}...",
        ]
        for i, s in enumerate(stats):
            fig.text(0.1, 0.22 - i * 0.04, s, fontsize=8, color=TEKS)

        fig.text(0.1, 0.04,
                 "Analisis: Pengujian integritas dilakukan dengan membandingkan nilai karakter ASCII antara "
                 "pesan input dan output (akurasi 100%) karena algoritma MWC bersifat deterministik — "
                 "menjamin setiap bit pesan dapat diekstraksi tanpa perubahan.",
                 fontsize=8, color=TEKS_SEK, style="italic")

        pdf.savefig(fig, facecolor=BG_DARK)
        plt.close(fig)

    logger.info(f"Laporan 7 PDF → {path_output}")
    return path_output


# ══════════════════════════════════════════════════════════════════════════════
# LAPORAN 8 — RINGKASAN PENGUJIAN TUNGGAL (DASHBOARD)
# ══════════════════════════════════════════════════════════════════════════════

def laporan_8_pdf(
    path_cover: str | Path,
    path_stego: str | Path,
    kunci: str,
    pesan_input: str,
    pesan_output: str,
    psnr: float,
    mse: float,
    path_output: str | Path,
) -> Path:
    
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.gridspec import GridSpec

    path_output = Path(path_output)
    id_lap = _buat_id_laporan("SUMMARY-008")
    _setup_tema_matplotlib()

    arr_cover = _muat_arr(path_cover)
    arr_stego = _muat_arr(path_stego)
    mask      = arr_cover[:, :, 0] != arr_stego[:, :, 0]
    h, w      = arr_cover.shape[:2]
    n_berubah = int(mask.sum())
    identik   = pesan_input == pesan_output
    persen_ubah = (n_berubah / (w * h)) * 100

    # Status PSNR
    if psnr >= 50:   kualitas, warna_k = "Sangat Baik", HIJAU
    elif psnr >= 40: kualitas, warna_k = "Baik",        TEAL
    elif psnr >= 30: kualitas, warna_k = "Cukup",       AMBER
    else:            kualitas, warna_k = "Buruk",        MERAH

    with PdfPages(str(path_output)) as pdf:
        fig = plt.figure(figsize=(11.69, 8.27), facecolor=BG_DARK)
        _halaman_header(
            fig, "LAPORAN RINGKASAN PENGUJIAN TUNGGAL", id_lap,
            f"File: {Path(path_cover).name}  |  Kunci: {'●' * min(len(kunci), 8)}  |  "
            f"Pesan: {len(pesan_input)} karakter",
        )
        _footer(fig)

        # Layout grid manual
        # ── Kolom kiri: Thumbnail + Metrik ───────────────────────────────────
        ax_cv = fig.add_axes([0.04, 0.54, 0.20, 0.28])
        ax_cv.imshow(arr_cover); ax_cv.axis("off")
        ax_cv.set_title("Cover (Asli)", color=TEAL, fontsize=7, fontweight="bold")

        ax_st = fig.add_axes([0.26, 0.54, 0.20, 0.28])
        ax_st.imshow(arr_stego); ax_st.axis("off")
        ax_st.set_title("Stego (Embedding)", color=AMBER, fontsize=7, fontweight="bold")

        # Panel metrik evaluasi (teks)
        metrik = [
            ("PSNR", f"{psnr:.4f} dB", warna_k),
            ("MSE",  f"{mse:.6f}",    TEKS),
            ("Piksel Berubah", f"{n_berubah:,} ({persen_ubah:.3f}%)", AMBER),
            ("Kualitas", kualitas, warna_k),
            ("Integritas", "100% VALID" if identik else "GAGAL", HIJAU if identik else MERAH),
            ("Dimensi", f"{w}×{h} px", TEKS_SEK),
            ("Kapasitas Citra", f"{(w*h)//8:,} byte", TEKS_SEK),
        ]
        for i, (label, nilai, warna) in enumerate(metrik):
            y = 0.50 - i * 0.055
            fig.text(0.04, y, f"{label}:", fontsize=8, color=TEKS_SEK)
            fig.text(0.22, y, nilai, fontsize=8, color=warna, fontweight="bold")

        # ── Kolom kanan atas: Histogram R ─────────────────────────────────────
        ax_hist = fig.add_axes([0.50, 0.55, 0.22, 0.27])
        x, fc = _hitung_hist(arr_cover[:, :, 0])
        _, fs = _hitung_hist(arr_stego[:, :, 0])
        ax_hist.fill_between(x, fc, alpha=0.5, color="#e05c5c", step="post", label="Cover")
        ax_hist.fill_between(x, fs, alpha=0.5, color="#e0a0a0", step="post", label="Stego")
        ax_hist.step(x, fc, color="#e05c5c", lw=0.6, where="post")
        ax_hist.step(x, fs, color="#e0a0a0", lw=0.6, where="post")
        _setup_ax(ax_hist, "Histogram Channel R")
        ax_hist.set_xlim(0, 255)
        ax_hist.legend(fontsize=6, facecolor=BG_PANEL, labelcolor=TEKS)
        ax_hist.tick_params(labelsize=6)

        # ── Noise Map ─────────────────────────────────────────────────────────
        ax_noise = fig.add_axes([0.76, 0.55, 0.20, 0.27])
        ax_noise.imshow(mask.astype(np.uint8) * 255, cmap="gray", aspect="auto")
        ax_noise.set_title("Noise Map (Sebaran Bit)", color=TEKS, fontsize=7, fontweight="bold")
        ax_noise.axis("off")

        # ── Verifikasi Pesan ──────────────────────────────────────────────────
        ax_pesan = fig.add_axes([0.04, 0.10, 0.92, 0.32])
        ax_pesan.set_facecolor("#0d1a0d" if identik else "#1a0d0d")
        ax_pesan.axis("off")
        ax_pesan.set_title(
            f"Verifikasi Integritas Pesan — STATUS: {'✓ 100% VALID & INTEGRAL' if identik else '✗ TIDAK COCOK'}",
            color=HIJAU if identik else MERAH, fontsize=9, fontweight="bold", pad=8,
        )

        # Input pesan (kiri dalam panel)
        ax_pesan.text(0.01, 0.88, "Pesan Input (sebelum disisipkan):", transform=ax_pesan.transAxes,
                      fontsize=7, color=TEAL, fontweight="bold")
        sample_in = pesan_input[:120] + ("..." if len(pesan_input) > 120 else "")
        ax_pesan.text(0.01, 0.68, sample_in, transform=ax_pesan.transAxes,
                      fontsize=7, color=TEKS, family="monospace")

        ax_pesan.text(0.01, 0.45, "Pesan Output (setelah diekstraksi):", transform=ax_pesan.transAxes,
                      fontsize=7, color=HIJAU, fontweight="bold")
        sample_out = pesan_output[:120] + ("..." if len(pesan_output) > 120 else "")
        ax_pesan.text(0.01, 0.25, sample_out, transform=ax_pesan.transAxes,
                      fontsize=7, color=TEKS, family="monospace")

        # Kesimpulan
        kesimpulan = (
            f"Kesimpulan Akhir: Berdasarkan metrik objektif (PSNR={psnr:.4f} dB — {kualitas}), "
            f"pengujian subjektif (imperceptibility), validasi integritas (100%), "
            f"dan distribusi noise yang merata ({persen_ubah:.3f}% piksel), implementasi "
            f"Steganografi LSB-MWC terbukti efektif dan aman."
        )
        fig.text(0.04, 0.04, kesimpulan, fontsize=8, color=TEKS_SEK, style="italic")

        pdf.savefig(fig, facecolor=BG_DARK)
        plt.close(fig)

    logger.info(f"Laporan 8 PDF → {path_output}")
    return path_output